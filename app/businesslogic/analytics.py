import os
import boto3
import json
import pandas as pd
import io
import sys
import time
import re
import openpyxl
import numpy as np

class analyticsClass():

    def __init__(self):
        self.bedrock = boto3.client(service_name='bedrock-runtime', region_name='us-west-2')

    def evaluate_formula(self, cell):
        try:
            value = cell.value
            return value if value is not None else cell.formula
        except Exception as e:
            print(f"Error evaluating formula: {e}")
            return None

    def exceltocsvParser(self, excel_file_path, sheet_name, output_file_path):
        wb = openpyxl.load_workbook(excel_file_path, data_only=True)
        ws = wb[sheet_name]

        # Evaluate and replace formulas with calculated values
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.data_type == 'f':
                    cell.value = self.evaluate_formula(cell)

        # Export the updated values to CSV
        with open(output_file_path, 'w') as f:
            for row in ws.iter_rows(values_only=True):
                f.write(','.join(str(v) for v in row) + '\n')


    inputkey = "Comparator_5_definition.maintenance_treatment"
    def create_qa_dataframe(self,inputkey, table_metadata, excel_file_path, output_file_path):
        # Extract values from table_metadata based on user input
        if inputkey in table_metadata:
            metadata_values = table_metadata[inputkey]
            
            # Extract main key values to a variable
            nrows = metadata_values["nrows"]
            skiprows = metadata_values["skiprows"]
            start_col = metadata_values["start_col"]
            end_col = metadata_values["end_col"]
            # Extract subpart before dot and remove underscores
            sheet_name = inputkey.split('.')[0].replace('_', ' ')

            self.exceltocsvParser(excel_file_path, sheet_name, output_file_path)

            csv_file_path = output_file_path
            df = pd.read_csv(csv_file_path, nrows=nrows, skiprows=skiprows, usecols=range(start_col, end_col))
            return df 
        else:
            return  "key not found"
        

    def genAIQA(self, df, prompt_input):
        try:
            column_names = df.columns.tolist()
            input_text=prompt_input
            words =  input_text.split() 
            if len(input_text) == 0 or input_text.replace(" ", "").isdigit() or all(
                    word.isdigit() or any(not char.isalpha() for char in word) for word in words):
                result='Please enter a valid input.'
            else:
                start_time = time.time()
                namespace = {}
                prompt = """Human: Please regard the following data:\n {}. Please provide Python code to Human Input "{}". If you're reading pandas dataframe please put low_memory=False and you always need to read csv as pd.read_csv.
                Print the data as per the requirement.Avoid including introductory lines or explanations in the python code and exclude user-specific comments or instructions such as paths or notes for replacement.
                Before Generating python code please cross verify all the code if any syntax errors please resolve before sending repsonse. 
                Python code should always be a single string with proper line breakers and comments its can't be a bullet points and it will start with import only.
                After importing the libraries please keep the code:
                pd.set_option('display.max_rows', None)  # Set the maximum number of rows to None for unlimited rows
                pd.set_option('display.max_columns', None)  # Set the maximum number of columns to None for unlimited columns
                Please exclude methods like 'describe()' and 'corr()' for columns with non-numeric data types
                In case any extra text came with python code please put those in a comment.
                This string can be really long please give in single string only, and dataframe variable always it a df only.
                While generating graphs and charts, please use Plotly Library only.
                Use object name as 'fig'  to represents a figure that can be rendered and displayed using Plotly. 
                If there are multiple plotly figure objects term it as 'fig1', 'fig2','fig3' and so on.
                Save the graph in a file with filename plotly_graph1.html using fig.write_html().
                If there are multiple plotly graphs then Save the graph in an multiple files with filenames plotly_graph1.html, plotly_graph2.html and so on, using fig.write_html().
                Check before writing plotly graph in files if that file exist or not, if file does not exist then create it using syntax open(file_name, "w", encoding="utf-8").close() where file_name can be plotly_graph.html, plotly_graph1.html etc.
                Import module "os" in python code.
                Specify a encoding when opening the file that can handle more Unicode characters like 'utf-8'
                In fig.write_html() under config attribute set displayModeBar: True 
                In fig.write_html() set auto_open=False and do not call fig.show()
                If python code contains df result as a table make sure to convert it to html using ".to_html(classes="table")" and print the result, For example : print(df.head().to_html(classes="table table-sm table-bordered tablecustom",  justify='left')).
                "please don't support Garbage text, Single word, and Meaningless sentences", and just print('Please enter valid Text').
                Assistant: 
                """.format(df.columns, input_text)
                if column_names != []:
                    print("exec")
                    body = json.dumps({
                        "prompt": prompt, 
                        "max_tokens_to_sample": 2048,
                        "temperature": 0.5,
                        "top_k": 250,
                        "top_p": 0.5,
                        })
                    modelId = 'anthropic.claude-v2'   
                    accept = 'application/json'
                    contentType = 'application/json'
                    response = self.bedrock.invoke_model(body=body, modelId=modelId, accept=accept, contentType=contentType)
                    response_body = json.loads(response.get('body').read())
                    result = response_body.get('completion')
                    print(result)
                    if 'python' in result.lower() or 'import' in result:
                        try:
                            code_start = result.find("```python") 
                            code_end = result.find("```", code_start + len("```python"))
                            python_code = result[code_start + len("```python") : code_end]
                            namespace = {'df': df}
                            python_code = python_code.replace('df = pd.read_csv','# df = pd.read_csv')

                            file_names = []
                            pattern = r'plotly_graph\d+.html'
                            pattern2 = r'plotly_graph.html'
                            matches = re.findall(pattern, result)
                            matches2 = re.findall(pattern2, result)
                            for match in matches:
                                file_names.append(match)
                            for match2 in matches2:
                                file_names.append(match2)
                            file_names = list(set(file_names))
                            print(file_names)

                            if len(file_names) > 0:
                                print("PLOTLY OUTPUT EXECUTED")
                                exec(python_code, namespace)
                                #APPEND HTML CODE IN RESULT FROM ALL FILES
                                result=''
                                for html_files in file_names:
                                    with open(html_files, "r", encoding="utf-8") as html_file: 
                                        html_code = html_file.read()
                                        result = result + html_code
                                #DELETE FILES
                                for html_file in file_names:
                                    os.remove(html_file)
                            else:
                                print("PLAIN TEXT OUTPUT EXECUTED")
                                # Execute the Python code and capture the output
                                old_stdout = sys.stdout
                                new_stdout = io.StringIO()
                                sys.stdout = new_stdout
                                exec(python_code, namespace) 
                                code_output = new_stdout.getvalue()
                                sys.stdout = old_stdout  
                                result=code_output
                                print(result)
                                
                        except Exception as e:
                            # data.initReq= True
                            result = 'Python Code Execution Error. Something went wrong while doing the analysis, please enter your input again.' + "ERROR:" + str(e)
            end_time = time.time()
            response_time = end_time - start_time
            print("Bedrock Response time :",response_time)
            return result 
        except Exception as e:
            error=str(e)  
            return "Exception in business logic. An issue has occurred. Please consider refreshing the page, or come back later."
        





